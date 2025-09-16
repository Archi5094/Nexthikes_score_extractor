import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

# ----------------------------
# CONFIG
# ----------------------------
LOGIN_URL = "https://digicrome.org/admin/login"
CHAT_URL_TEMPLATE = "https://digicrome.org/admin/internship/singlechat?search={email}"

USERNAME = "twinklebaid@nexthikes.com"
PASSWORD = "manila"
st.set_page_config(
    page_title="Nexthikes Score Recorder",
    page_icon="📊",
    layout="centered"
)

# ----------------------------
# LOGIN FUNCTION
# ----------------------------
def login_to_portal(username, password):
    s = requests.Session()
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
        "Referer": LOGIN_URL,
        "Accept-Language": "en-US,en;q=0.9",
    }

    r = s.get(LOGIN_URL, headers=headers, timeout=20)
    soup = BeautifulSoup(r.text, "html.parser")
    token_input = soup.find("input", {"name": "_token"})
    if not token_input:
        return None
    csrf_token = token_input["value"]

    data = {"email": username, "password": password, "_token": csrf_token}
    resp = s.post(LOGIN_URL, data=data, headers=headers, timeout=20, allow_redirects=True)

    if "logout" in resp.text.lower() or "dashboard" in resp.text.lower():
        return s
    return None


# ----------------------------
# FETCH CHAT HISTORY
# ----------------------------
def get_chat_history(session, email):
    chat_url = CHAT_URL_TEMPLATE.format(email=email)
    resp = session.get(chat_url, timeout=20)
    soup = BeautifulSoup(resp.text, "html.parser")
    chat_screen = soup.find("div", id="chat-screen")

    if not chat_screen:
        return ""

    messages = []
    for parent_div in chat_screen.find_all("div", style=lambda v: v and "text-align" in v):
        align = parent_div.get("style", "")
        if "left" in align:  # student messages
            msg_div = parent_div.find("div", class_="alert")
            if not msg_div:
                continue
            text = msg_div.get_text(" ", strip=True)
            time_tag = msg_div.find("small")
            timestamp = time_tag.get_text(strip=True) if time_tag else None
            if timestamp:
                text = text.replace(timestamp, "").strip()
                messages.append(f"[{timestamp}] {text}")
            else:
                messages.append(text)

    return "\n".join(messages)

# ----------------------------
# EXTRACT FEEDBACK
# ----------------------------
def extract_feedback(chat_history, project_range):
    results = {}
    for i in project_range:
        project_key = f"Project {i}"
        score, grade, feedback = "Not Submitted", "Not Submitted", "Not Submitted"

        if chat_history.strip():
            score_match = re.search(rf"Project {i}.*?Score[:\s]*([0-9]+)", chat_history, re.IGNORECASE | re.DOTALL)
            grade_match = re.search(rf"Project {i}.*?Grade[:\s]*([A-F][+-]?)", chat_history, re.IGNORECASE | re.DOTALL)
            feedback_match = re.search(rf"Project {i}.*?Feedback[:\s]*(.+?)(?:Project \d+|$)", chat_history, re.IGNORECASE | re.DOTALL)

            if score_match: score = score_match.group(1).strip()
            if grade_match: grade = grade_match.group(1).strip()
            if feedback_match: feedback = feedback_match.group(1).strip()

        results[project_key] = {"score": score, "grade": grade, "feedback": feedback}
    return results

# ----------------------------
# EXPORT TO EXCEL (STYLED)
# ----------------------------
def export_feedback_to_excel(feedback_all, project_range):
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback"

    # Header row
    header = ["Email"] + [f"Project {i}" for i in project_range]
    ws.append(header)

    # Style header
    header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for email, projects in feedback_all.items():
        row = [email]
        for i in project_range:
            project_key = f"Project {i}"
            score = projects.get(project_key, {}).get("score", "Not Submitted")
            row.append(score)
        ws.append(row)

        # Add grade + feedback as comment
        for col_idx, i in enumerate(project_range, start=2):
            project_key = f"Project {i}"
            if project_key in projects:
                cell = ws.cell(row=ws.max_row, column=col_idx)
                grade = projects[project_key]["grade"]
                feedback = projects[project_key]["feedback"]
                cell.comment = Comment(f"Grade: {grade}\nFeedback:\n{feedback}", "Nexthikes")

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ----------------------------
# STREAMLIT APP
# ----------------------------
st.markdown(
    """
    <div style="text-align:center; padding:10px;">
        <h2 style="color:#003366;">📊 Nexthikes Score Recorder</h2>
        <p style="color:#666;">Extract internship project scores & feedback directly from chat logs.</p>
    </div>
    """,
    unsafe_allow_html=True
)

# Sidebar credentials
with st.sidebar:
    st.header("🔑 Admin Login")
    username = st.text_input("Admin Email")
    password = st.text_input("Password", type="password")

# Main inputs
st.subheader("⚙️ Extraction Settings")
batch_name = st.text_input("📌 Enter Batch Name", placeholder="e.g. DataScience_July2025")
uploaded_file = st.file_uploader("📂 Upload Excel with Emails", type=["xlsx"])

col1, col2 = st.columns(2)
with col1:
    start_proj = st.number_input("Start Project Number", min_value=1, max_value=20, value=1)
with col2:
    end_proj = st.number_input("End Project Number", min_value=1, max_value=20, value=4)
project_range = range(start_proj, end_proj + 1)

# Run extraction
if st.button("🚀 Run Extraction"):
    if not uploaded_file or not username or not password or not batch_name:
        st.error("⚠️ Please fill all fields and upload the Excel file.")
    else:
        df = pd.read_excel(uploaded_file)
        if "email" not in df.columns:
            st.error("Excel must have a column named 'email'")
        else:
            st.info("🔄 Logging in to portal...")
            session = login_to_portal(username,password)
            if not session:
                st.error("❌ Login failed. Check credentials.")
            else:
                st.success("✅ Login successful!")

                feedback_all = {}
                progress = st.progress(0)
                for idx, row in df.iterrows():
                    email = row["email"]
                    st.write(f"📩 Processing: {email}")
                    chat_history = get_chat_history(session, email)

                    if not chat_history.strip():
                        feedback = {f"Project {i}": {"score": "Not Submitted", "grade": "Not Submitted", "feedback": "Not Submitted"} for i in project_range}
                    else:
                        feedback = extract_feedback(chat_history, project_range)

                    feedback_all[email] = feedback
                    progress.progress((idx + 1) / len(df))

                st.success("🎉 Extraction complete!")

                file_name = f"{batch_name}_Project{start_proj}-{end_proj}_scores.xlsx"
                excel_file = export_feedback_to_excel(feedback_all, project_range)
                st.download_button(
                    label=f"📥 Download {file_name}",
                    data=excel_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )




